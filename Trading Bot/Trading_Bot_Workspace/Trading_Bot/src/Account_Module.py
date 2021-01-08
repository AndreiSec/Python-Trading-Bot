'''
Created on Sep 16, 2020

@author: Andrei Secara

Module designed to create and manage bank accounts.
'''
import openpyxl
from datetime import date
from Stock_Module import single_stock_price_checker_loop
#Account 1 for testing:
# account_1_name = "Account 1"
# account_1_balance = 3000
# account_1_risk_tolerance = 0.5


account_1 = "C://Users//andrei//Desktop//Trading Bot//Accounts//Account_1.xlsx"

#Define account as an object for easier use of attributes.
# class Account:
#     def __init__(self):
#         self.name = ""
#         self.balance = 0
#         self.risk_tolerance = 0
# 
# #For use when user input wanted
# def account_specifier_input():
#     account = Account()
#     account.name = str(input("Enter an account name: "))
#     account.balance = float(input("Enter account starting balance: "))
#     account.risk_tolerance = float(input("Enter account risk tolerance (between 0 and 1). 0 Being non-existent risk, 1 being high risk. "))
#     return account
# 
# #For use during testing
# def account_specifier_defined():
#     account = Account()
#     account.name = str(account_1_name)
#     account.balance = float(account_1_balance)
#     account.risk_tolerance = float(account_1_risk_tolerance)
#     return account

def account_balance_retriever(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
#     print(ws.max_row)
    balance = float(ws.cell(column=2, row=ws.max_row).value)
    
    return balance


#Updates excel file account balance
def account_balance_updater(filepath, amount):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    ws.cell(column=1, row=ws.max_row+1, value=date.today())
    ws.cell(column=2, row=ws.max_row, value=amount)
    wb.save(filepath)
    return

def buy_recorder(filepath, stock_ticker, price, amount):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    ws.cell(column=3, row=ws.max_row, value= str(str(stock_ticker) + " " + str(price) + " " + str(amount)))
    
    wb.save(filepath)
    return
    
def sell_recorder(filepath, stock_ticker, price, amount):
    
    temp_acc_balance = account_balance_retriever(filepath)
    
    
    
    account_balance_updater(filepath, temp_acc_balance+(price*amount))
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    ws.cell(column=4, row=ws.max_row, value= str(str(stock_ticker + " " + str(price) + " " + str(amount))))
    
    wb.save(filepath)
    return

def buy_and_sell(stock_ticker, price):
    balance = account_balance_retriever(account_1)
#     print(balance)
    if price > 0:
        amount_to_buy = balance // price
    else:
        amount_to_buy = 0
    bought_complete = False
    if amount_to_buy >= 1:
        balance = balance - amount_to_buy*price
        account_balance_updater(account_1, balance)
        buy_recorder(account_1, stock_ticker, price, amount_to_buy)
    
        price_to_sell = single_stock_price_checker_loop(price, stock_ticker)
    
        sell_recorder(account_1, stock_ticker, price_to_sell, amount_to_buy)
        bought_complete = True
    
    return bought_complete


# balance = account_balance_retriever(account_1)
# print(balance)
# price = 108.86
# amount_to_buy = balance // price
# balance = balance - amount_to_buy*price
# account_balance_updater(account_1, balance)
# buy_recorder(account_1, "AAPL", price, amount_to_buy)
# sell_recorder(account_1, "AAPL", 110.32, amount_to_buy)


# filepath = "C://Users//andrei//Desktop//Trading Bot//Accounts//Account_1.xlsx"
# amount = 10
# print(account_balance_retriever(filepath))