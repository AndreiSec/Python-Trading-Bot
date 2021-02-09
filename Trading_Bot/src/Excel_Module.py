'''
Created on Sep 16, 2020

@author: Andrei Secara

Module to load in excel data
'''
# import pandas as pd
import openpyxl



def create_stock_list(stock_list_file_name, stock_list_sheet):
#     excel_data = pd.read_excel(stock_list_file_name, sheet_name= stock_list_sheet)
    stock_list = []
    wb = openpyxl.load_workbook(filename=stock_list_file_name)
    ws = wb[stock_list_sheet]
#     for cell in ws.columns['A']:
#         stock_list.append(cell.value)
#     print(str(stock_list_file_name) + " max row =  " + str(ws.max_row))
    for i in range(1, ws.max_row):
        
#         print(ws.cell(row = i, column = 1).value)
        stock_list.append(str(ws.cell(row = i, column = 1).value))
        
#     stock_list.sort()
#     print(stock_list)
    wb.close()
    return stock_list
    
def create_price_list(stock_list_file_name, stock_list_sheet):
    price_list = []
    wb = openpyxl.load_workbook(filename=stock_list_file_name)
    ws = wb[stock_list_sheet]
#     for cell in ws.columns['A']:
#         stock_list.append(cell.value)
    for i in range(1, ws.max_row):
#         print(ws.cell(row = i, column = 1).value)
        price_list.append(float(ws.cell(row = i, column = 2).value))
        
#     stock_list.sort()
#     print(stock_list)
    wb.close()
    return price_list
    